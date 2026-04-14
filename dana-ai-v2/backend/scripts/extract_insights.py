import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl, re, json, os
import numpy as np

INSIGHT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'insight.xlsx')
ER_OUT_PATH      = os.path.join(os.path.dirname(__file__), '..', 'data', 'er_data.json')
PATTERNS_OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'campaign_patterns.json')


# ── Tier classification dari followers ─────────────────────────
def classify_tier(followers):
    if not followers or followers == 0:
        return 'unknown'
    if followers < 10_000:
        return 'nano'
    if followers < 100_000:
        return 'mikro'
    if followers < 1_000_000:
        return 'makro'
    return 'mega'


def classify_er_quality(er_pct):
    """Klasifikasi kualitas ER"""
    if er_pct is None:
        return 'unknown'
    if er_pct >= 15:
        return 'excellent'
    if er_pct >= 8:
        return 'good'
    if er_pct >= 3:
        return 'average'
    return 'low'


def extract():
    wb = openpyxl.load_workbook(INSIGHT_PATH, data_only=True)
    
    # ── Per-KOL records ──────────────────────────────────────────
    all_kol_records = {}  # username -> list of records
    
    # ── Campaign-level records ───────────────────────────────────
    # Berisi semua data mentah dari setiap KOL di setiap campaign
    all_records = []
    
    # ── Sheet3: individual KOL ER tracking ─────────────────────
    if 'Sheet3' in wb.sheetnames:
        ws = wb['Sheet3']
        headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows())]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1] or not isinstance(row[1], str):
                continue
            uname = row[1].strip().lower().replace(' ', '').replace('@', '')
            if uname in ('accountusername', '', 'total') or 'total' in uname.lower():
                continue

            followers  = row[2] if isinstance(row[2], (int, float)) and row[2] > 0 else 0
            reach      = row[4] if isinstance(row[4], (int, float)) and row[4] > 0 else 0
            likes      = row[5] if isinstance(row[5], (int, float)) else 0
            comments   = row[6] if isinstance(row[6], (int, float)) else 0
            shares     = row[7] if isinstance(row[7], (int, float)) else 0
            saves      = row[8] if isinstance(row[8], (int, float)) else 0
            total_eng  = row[12] if isinstance(row[12], (int, float)) else 0
            er         = row[13] if isinstance(row[13], (int, float)) and 0 < row[13] < 200 else None

            record = {
                'username':  uname,
                'followers': followers,
                'reach':     reach,
                'er_pct':    er,
                'total_eng': total_eng,
                'tier':      classify_tier(followers),
                'er_quality': classify_er_quality(er),
                'reach_ratio': round(reach / followers, 4) if followers > 0 else 0,
            }
            
            if uname not in all_kol_records:
                all_kol_records[uname] = []
            all_kol_records[uname].append(record)
            all_records.append(record)

    # ── Sheet4: campaign KOL performance ─────────────────────
    if 'Sheet4' in wb.sheetnames:
        ws = wb['Sheet4']
        for row in ws.iter_rows(values_only=True):
            p = row[0]
            if not p or not isinstance(p, str):
                continue
            m = re.search(r'KOL\s*\d+\s*[-]\s*(?:@)?([\w.]+)', p, re.IGNORECASE)
            if not m:
                continue
            uname = m.group(1).lower().replace('@', '')

            followers  = row[3] if isinstance(row[3], (int, float)) and row[3] > 0 else 0
            reach      = row[6] if isinstance(row[6], (int, float)) and row[6] > 0 else 0
            total_eng  = row[14] if isinstance(row[14], (int, float)) else 0
            er         = row[15] if isinstance(row[15], (int, float)) and 0 < row[15] < 200 else None
            
            record = {
                'username':   uname,
                'followers':  followers,
                'reach':      reach,
                'er_pct':     er,
                'total_eng':  total_eng,
                'tier':       classify_tier(followers),
                'er_quality': classify_er_quality(er),
                'reach_ratio': round(reach / followers, 4) if followers > 0 else 0,
            }
            
            if uname not in all_kol_records:
                all_kol_records[uname] = []
            all_kol_records[uname].append(record)
            all_records.append(record)

    # ── Aggregate per-KOL ────────────────────────────────────────
    er_summary = {}
    for uname, records in all_kol_records.items():
        if uname == 'accountusername':
            continue
        valid_er   = [r['er_pct'] for r in records if r['er_pct'] is not None]
        avg_er     = sum(valid_er) / len(valid_er) if valid_er else None
        avg_reach  = sum(r['reach'] for r in records) / len(records)
        avg_followers = sum(r['followers'] for r in records) / len(records) if any(r['followers'] for r in records) else 0
        
        er_summary[uname] = {
            'avg_er_pct':   round(avg_er, 2) if avg_er is not None else None,
            'avg_reach':    round(avg_reach),
            'avg_followers': round(avg_followers),
            'post_count':   len(records),
            'tier':         classify_tier(avg_followers),
            'er_quality':   classify_er_quality(avg_er),
        }

    # ── Simpan ER data (backward compatible) ─────────────────────
    er_out = {u: {
        'avg_er_pct':  v['avg_er_pct'],
        'avg_reach':   v['avg_reach'],
        'post_count':  v['post_count'],
    } for u, v in er_summary.items()}
    
    with open(ER_OUT_PATH, 'w') as f:
        json.dump(er_out, f, indent=2)
    print(f"[OK] ER data: {len(er_summary)} KOL")

    # ── Analisis campaign patterns ────────────────────────────────
    patterns = analyze_patterns(all_records, er_summary)
    
    with open(PATTERNS_OUT_PATH, 'w') as f:
        json.dump(patterns, f, indent=2)
    print(f"[OK] Campaign patterns saved -> {PATTERNS_OUT_PATH}")
    
    return er_summary, patterns


def analyze_patterns(all_records, er_summary):
    """
    Analisis pola dari semua campaign insight:
    - Tier apa yang paling sering perform bagus
    - Followers range optimal
    - ER distribution per tier
    - Reach ratio pattern
    """
    if not all_records:
        return {}

    valid_records = [r for r in all_records if r['er_pct'] is not None]
    
    if not valid_records:
        return {'note': 'Tidak ada data ER yang valid di insight'}

    # ── Per-tier analysis ─────────────────────────────────────────
    tier_stats = {}
    for tier in ['nano', 'mikro', 'makro', 'mega']:
        tier_recs = [r for r in valid_records if r['tier'] == tier]
        if not tier_recs:
            continue
        ers = [r['er_pct'] for r in tier_recs]
        reaches = [r['reach'] for r in tier_recs if r['reach'] > 0]
        
        tier_stats[tier] = {
            'count':        len(tier_recs),
            'avg_er':       round(np.mean(ers), 2),
            'median_er':    round(float(np.median(ers)), 2),
            'max_er':       round(max(ers), 2),
            'min_er':       round(min(ers), 2),
            'avg_reach':    round(np.mean(reaches)) if reaches else 0,
            'pct_excellent': round(len([e for e in ers if e >= 15]) / len(ers) * 100, 1),
            'pct_good':      round(len([e for e in ers if e >= 8]) / len(ers) * 100, 1),
        }

    # ── Best performing tier ──────────────────────────────────────
    best_tier = None
    best_er   = 0
    for tier, stats in tier_stats.items():
        if stats['avg_er'] > best_er:
            best_er   = stats['avg_er']
            best_tier = tier

    # ── ER quality distribution ───────────────────────────────────
    quality_dist = {}
    for q in ['excellent', 'good', 'average', 'low']:
        cnt = len([r for r in valid_records if r['er_quality'] == q])
        quality_dist[q] = {
            'count': cnt,
            'pct':   round(cnt / len(valid_records) * 100, 1)
        }

    # ── Followers range yang optimal (top 25% ER) ─────────────────
    sorted_by_er  = sorted(valid_records, key=lambda x: x['er_pct'], reverse=True)
    top_quartile  = sorted_by_er[:max(1, len(sorted_by_er)//4)]
    top_followers = [r['followers'] for r in top_quartile if r['followers'] > 0]
    
    optimal_followers = {}
    if top_followers:
        optimal_followers = {
            'min': int(min(top_followers)),
            'max': int(max(top_followers)),
            'avg': int(np.mean(top_followers)),
            'median': int(np.median(top_followers)),
        }

    # ── Reach ratio patterns ──────────────────────────────────────
    # reach_ratio = reach/followers — makin tinggi makin baik (konten nyebar)
    reach_ratios = [r['reach_ratio'] for r in valid_records if r['reach_ratio'] > 0]
    reach_stats  = {}
    if reach_ratios:
        reach_stats = {
            'avg':    round(float(np.mean(reach_ratios)), 4),
            'median': round(float(np.median(reach_ratios)), 4),
            'p75':    round(float(np.percentile(reach_ratios, 75)), 4),
            'p90':    round(float(np.percentile(reach_ratios, 90)), 4),
        }

    # ── Score weight recommendations ─────────────────────────────
    # Berdasarkan data insight, hitung bobot rekomendasi
    # Kalau nano perform jauh lebih baik -> naikkan bobot tier nano
    tier_weight_hint = {}
    if tier_stats:
        all_avg = np.mean([s['avg_er'] for s in tier_stats.values()])
        for tier, stats in tier_stats.items():
            ratio = stats['avg_er'] / all_avg if all_avg > 0 else 1.0
            tier_weight_hint[tier] = round(float(ratio), 3)

    # ── Summary untuk display di UI ──────────────────────────────
    all_ers = [r['er_pct'] for r in valid_records]
    overall_avg_er = round(float(np.mean(all_ers)), 2)
    
    patterns = {
        'total_records':        len(all_records),
        'total_with_er':        len(valid_records),
        'overall_avg_er':       overall_avg_er,
        'overall_median_er':    round(float(np.median(all_ers)), 2),
        'best_tier':            best_tier,
        'best_tier_avg_er':     best_er,
        'tier_stats':           tier_stats,
        'quality_distribution': quality_dist,
        'optimal_followers':    optimal_followers,
        'reach_ratio_stats':    reach_stats,
        'tier_weight_hint':     tier_weight_hint,
        'insights_summary': {
            'recommendation': f"Berdasarkan {len(valid_records)} data campaign, tier {best_tier} perform paling baik dengan rata-rata ER {best_er:.1f}%",
            'top_er_range': f"ER range dari dataset: {min(all_ers):.1f}% - {max(all_ers):.1f}%",
        }
    }

    # Print summary
    print(f"\n[PATTERNS] Campaign Performance Analysis:")
    print(f"  Total records: {len(all_records)} | Dengan ER: {len(valid_records)}")
    print(f"  Overall avg ER: {overall_avg_er}%")
    print(f"  Best performing tier: {best_tier} (avg ER: {best_er:.1f}%)")
    print(f"\n  Per-tier breakdown:")
    for tier, stats in tier_stats.items():
        print(f"    {tier:8s}: n={stats['count']:3d} | avg ER={stats['avg_er']:5.1f}% | excellent={stats['pct_excellent']}%")
    if tier_weight_hint:
        print(f"\n  Tier weight hints (relatif thd rata-rata):")
        for tier, w in tier_weight_hint.items():
            bar = '█' * int(w * 10)
            print(f"    {tier:8s}: {w:.3f} {bar}")
    
    return patterns


if __name__ == '__main__':
    er_summary, patterns = extract()
    print(f"\n[DONE] er_data.json + campaign_patterns.json tersimpan")