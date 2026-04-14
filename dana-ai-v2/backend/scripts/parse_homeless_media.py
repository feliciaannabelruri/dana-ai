import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import pandas as pd
import numpy as np
import re, json, os

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
OUT_PATH = os.path.join(DATA_DIR, 'homeless_media.json')

HOMELESS_MEDIA_PATH = os.environ.get(
    'HOMELESS_MEDIA_PATH',
    os.path.join(DATA_DIR, 'HomelessMedia.xlsx')
)

LOCATION_GROUPS = {
    "jakarta":      ["jakarta","jaksel","jakpus","jakbar","jaktim","jakut","jkt",
                     "gading serpong","tangerang","depok","bekasi","bogor","bsd",
                     "serpong","cibubur","cikarang"],
    "bandung":      ["bandung","cimahi"],
    "cirebon":      ["cirebon"],
    "surabaya":     ["surabaya","sidoarjo","pasuruan","kediri","gresik"],
    "malang":       ["malang","batu"],
    "jawa_timur":   ["jawa timur","jatim","banyuwangi","jember","madiun",
                     "lumajang","blitar","mojokerto","probolinggo","lamongan",
                     "tuban","bojonegoro"],
    "yogyakarta":   ["yogyakarta","jogja","sleman","bantul","gunung kidul","kulon progo"],
    "solo":         ["solo","surakarta","karanganyar","wonogiri","klaten","boyolali","sragen"],
    "semarang":     ["semarang","salatiga","kendal","demak","ungaran"],
    "jawa_tengah":  ["jawa tengah","jateng","magelang","purwokerto","cilacap",
                     "banyumas","kebumen","wonosobo","temanggung","kudus",
                     "pati","jepara","rembang","blora","batang","pemalang",
                     "tegal","brebes","pekalongan","purbalingga","banjarnegara","grobogan"],
    "bali":         ["bali","denpasar","badung","gianyar","tabanan","buleleng",
                     "karangasem","klungkung","bangli","jembrana"],
    "sumatra":      ["medan","palembang","pekanbaru","pekan baru","batam","lampung",
                     "padang","aceh","jambi","bengkulu","banda aceh","langsa",
                     "lhokseumawe","binjai","pematangsiantar","lubuklinggau",
                     "prabumulih","dumai","padang sidempuan"],
    "kalimantan":   ["kalimantan","banjarmasin","samarinda","pontianak","balikpapan",
                     "palangkaraya","banjarbaru","tarakan","singkawang","kotabaru"],
    "sulawesi":     ["sulawesi","makassar","manado","gowa","palu","kendari",
                     "gorontalo","mamuju","palopo"],
    "nasional":     ["nasional","national","indonesia"],
}


def parse_followers(val):
    if val is None: return 0
    if isinstance(val, float) and np.isnan(val): return 0
    s = str(val).strip().upper().replace(',', '.').replace(' ', '')
    try:
        if 'M' in s: return int(float(re.sub(r'[^0-9.]', '', s)) * 1_000_000)
        if 'K' in s: return int(float(re.sub(r'[^0-9.]', '', s)) * 1_000)
        cleaned = re.sub(r'[^0-9.]', '', s)
        if not cleaned: return 0
        return int(float(cleaned))
    except:
        return 0


def parse_rate(val):
    """
    Robust rate parser. Rate values in the file are already full rupiah.
    Handles: int/float, '-', '27-28Jt', 'Rp3.500.00', 'Rp 3.500.000', etc.
    Returns 0 if unparseable.
    """
    if val is None: return 0
    if isinstance(val, (int, float)):
        try:
            v = int(val)
            return v if v > 0 else 0
        except:
            return 0

    s = str(val).strip()
    if s in ('', '-', 'nan', 'None', '#DIV/0!', 'N/A', 'n/a'):
        return 0

    s_lower = s.lower()

    # Handle range like '27-28Jt' → take the lower bound
    # Pattern: number-numberUnit or number-number
    range_match = re.match(r'^([0-9.,]+)\s*[-–]\s*([0-9.,]+)\s*(jt|juta|rb|ribu|k|m)?$', s_lower)
    if range_match:
        try:
            low_str = range_match.group(1).replace(',', '.')
            unit = range_match.group(3) or ''
            low = float(low_str)
            if unit in ('jt', 'juta'):
                low *= 1_000_000
            elif unit in ('rb', 'ribu', 'k'):
                low *= 1_000
            elif unit == 'm':
                low *= 1_000_000
            return int(low)
        except:
            pass

    # Handle 'Rp' prefix, dots as thousands separator, comma as decimal
    # e.g. 'Rp3.500.00' → could be 3500 (typo) or 350000
    s_clean = re.sub(r'[Rr][Pp]\.?\s*', '', s)  # remove Rp prefix

    # Check for unit suffix
    unit_match = re.search(r'(jt|juta|rb|ribu)$', s_clean.lower())
    unit_mult = 1
    if unit_match:
        u = unit_match.group(1)
        if u in ('jt', 'juta'):
            unit_mult = 1_000_000
        elif u in ('rb', 'ribu'):
            unit_mult = 1_000
        s_clean = s_clean[:unit_match.start()].strip()

    # Remove non-numeric except dots and commas
    # If has dots as thousands: 3.500.000 → 3500000
    # If has comma as decimal: 3,5 → 3.5
    s_clean = s_clean.strip().replace(' ', '')

    # Count dots and commas
    dot_count = s_clean.count('.')
    comma_count = s_clean.count(',')

    try:
        if dot_count > 1:
            # Multiple dots → thousands separator (e.g. 3.500.000)
            numeric = s_clean.replace('.', '').replace(',', '')
            return int(float(numeric)) * unit_mult
        elif dot_count == 1 and comma_count == 0:
            # Single dot — could be decimal or thousands
            parts = s_clean.split('.')
            if len(parts[1]) <= 2 and int(parts[1]) == 0:
                # Likely typo like '3.500.00' reduced → treat whole part as value
                return int(parts[0]) * unit_mult
            elif len(parts[1]) == 3:
                # Thousands separator (e.g. 3.500)
                return int(s_clean.replace('.', '')) * unit_mult
            else:
                return int(float(s_clean)) * unit_mult
        elif comma_count == 1 and dot_count == 0:
            # Comma as decimal or thousands
            parts = s_clean.split(',')
            if len(parts[1]) == 3:
                return int(s_clean.replace(',', '')) * unit_mult
            else:
                return int(float(s_clean.replace(',', '.'))) * unit_mult
        else:
            # Just digits
            numeric = re.sub(r'[^0-9]', '', s_clean)
            return int(numeric) * unit_mult if numeric else 0
    except:
        return 0


def parse_contact(val):
    if val is None: return ''
    if isinstance(val, float):
        if np.isnan(val): return ''
        try: return str(int(val)).strip()
        except: return ''
    if isinstance(val, int):
        return str(val).strip()
    return str(val).strip()


def normalize_location(loc):
    if not loc or str(loc).strip().lower() in ('nan', ''):
        return "nasional"
    loc_lower = str(loc).lower().strip()
    for group, keywords in LOCATION_GROUPS.items():
        if any(kw in loc_lower for kw in keywords):
            return group
    return "other"


def build_contact_action(contact_raw, username, social_media):
    contact = contact_raw.replace('+', '').replace(' ', '').replace('.0', '').strip()
    if contact and contact.isdigit() and len(contact) >= 8:
        if contact.startswith('0'):
            contact = '62' + contact[1:]
        elif not contact.startswith('62'):
            contact = '62' + contact
        return {
            'type': 'whatsapp',
            'url': f'https://wa.me/{contact}',
            'label': f'WA {contact}',
        }
    uname = username.lstrip('@')
    sm = str(social_media).lower()
    if 'tiktok' in sm:
        return {'type': 'tiktok', 'url': f'https://tiktok.com/@{uname}', 'label': f'DM @{uname}'}
    return {'type': 'instagram', 'url': f'https://instagram.com/{uname}', 'label': f'DM @{uname}'}


def parse_sheet(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)

    sheet_name = 'Sheet1' if 'Sheet1' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f"[*] Membaca sheet: '{sheet_name}'")

    records = []
    current = None
    last_category = 'Media'  # fallback category

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * max(0, 10 - len(row))

        no_val = row[0]
        try:
            has_id = no_val is not None and str(no_val).replace('.0', '').strip().isdigit()
        except:
            has_id = False

        if has_id:
            if current and current.get('username', '').strip():
                records.append(current)

            # Category: use col 4, fallback to last known if None/empty
            cat_raw = row[4]
            if cat_raw is not None and str(cat_raw).strip() not in ('', 'None', 'nan'):
                last_category = str(cat_raw).strip()
            category = last_category

            rate_platform = str(row[8]).strip() if row[8] is not None else ''
            rate_value = parse_rate(row[9])
            contact_raw = parse_contact(row[6])

            current = {
                'id':            int(float(no_val)),
                'username':      str(row[1]).strip() if row[1] is not None else '',
                'social_media':  str(row[2]).strip() if row[2] is not None else 'Instagram',
                'followers_raw': row[3],
                'category':      category,
                'pic_name':      str(row[5]).strip() if row[5] is not None else '',
                'pic_contact':   contact_raw,
                'location_raw':  str(row[7]).strip() if row[7] is not None else 'Nasional',
                'rate_card':     {},
            }
            if rate_platform and rate_value > 0:
                current['rate_card'][rate_platform] = rate_value

        elif current is not None:
            rate_platform = str(row[8]).strip() if row[8] is not None else ''
            rate_value = parse_rate(row[9])
            if rate_platform and rate_value > 0:
                current['rate_card'][rate_platform] = rate_value

    if current and current.get('username', '').strip():
        records.append(current)

    return records


def enrich(records):
    result = []
    for r in records:
        followers_num = parse_followers(r['followers_raw'])
        rates = [v for v in r['rate_card'].values() if v > 0]
        rate_min = min(rates) if rates else 0
        rate_max = max(rates) if rates else 0
        loc_norm = normalize_location(r['location_raw'])
        contact_action = build_contact_action(r['pic_contact'], r['username'], r['social_media'])

        result.append({
            'id':              r['id'],
            'username':        r['username'],
            'social_media':    r['social_media'],
            'followers_raw':   str(r['followers_raw']) if r['followers_raw'] else '0',
            'followers_num':   followers_num,
            'category':        r['category'],
            'pic_name':        r['pic_name'],
            'pic_contact':     r['pic_contact'],
            'location_raw':    r['location_raw'],
            'location_norm':   loc_norm,
            'rate_card':       r['rate_card'],
            'rate_min':        rate_min,
            'rate_max':        rate_max,
            'contact_action':  contact_action,
        })
    return result


def main():
    if not os.path.exists(HOMELESS_MEDIA_PATH):
        print(f"[ERROR] File tidak ditemukan: {HOMELESS_MEDIA_PATH}")
        sys.exit(1)

    print(f"[*] Parsing: {HOMELESS_MEDIA_PATH}")
    records = parse_sheet(HOMELESS_MEDIA_PATH)
    enriched = enrich(records)
    print(f"[OK] {len(enriched)} Homeless Media accounts parsed")

    # Sample output
    print("\n  Sample rate card (first 5):")
    for r in enriched[:5]:
        print(f"    @{r['username']}: cat={r['category']} | {r['rate_card']} | min={r['rate_min']} max={r['rate_max']}")

    out_path = OUT_PATH
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(enriched, f, ensure_ascii=False, indent=2)
    print(f"\n[OK] Saved → {out_path}")

    cats = {}
    locs = {}
    for r in enriched:
        cats[r['category']] = cats.get(r['category'], 0) + 1
        locs[r['location_norm']] = locs.get(r['location_norm'], 0) + 1

    print("\nKategori:")
    for cat, cnt in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {cnt}")

    print("\nLokasi (normalized):")
    for loc, cnt in sorted(locs.items(), key=lambda x: -x[1]):
        print(f"  {loc}: {cnt}")


if __name__ == '__main__':
    main()