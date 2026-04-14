import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import re, json, os
import numpy as np
import openpyxl

DATA_DIR    = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
KOL_FREE_IN = os.environ.get('KOL_HOMELESS_PATH', os.path.join(DATA_DIR, 'KOLHomeless.xlsx'))
COMM_IN     = os.environ.get('COMMUNITY_PATH',    os.path.join(DATA_DIR, 'Community.xlsx'))
KOL_FREE_OUT = os.path.join(DATA_DIR, 'kol_homeless_free.json')
COMM_OUT     = os.path.join(DATA_DIR, 'community_pool.json')


def parse_followers(val):
    if val is None: return 0
    if isinstance(val, float) and np.isnan(val): return 0
    s = str(val).strip().upper().replace(',', '.').replace(' ', '')
    try:
        if 'M' in s: return int(float(re.sub(r'[^0-9.]', '', s)) * 1_000_000)
        if 'K' in s: return int(float(re.sub(r'[^0-9.]', '', s)) * 1_000)
        cleaned = re.sub(r'[^0-9.]', '', s)
        return int(float(cleaned)) if cleaned else 0
    except:
        return 0


def approachability_score(contact_val):
    """
    Skor 0–1 seberapa mudah di-approach.
      dm / DM                → 1.0   (paling gampang)
      nomor WA (628xxx)      → 0.9
      email personal/brand   → 0.7
      email institusi formal → 0.5
      kosong                 → 0.3
    """
    if contact_val is None:
        return 0.3
    c = str(contact_val).strip().lower()
    if not c or c in ('nan', 'none', '-'):
        return 0.3
    if c == 'dm' or c.startswith('dm'):
        return 1.0
    # WA number
    if re.match(r'^(62|0|\+62)\d{8,}', c.replace(' ', '')):
        return 0.9
    # email
    if '@' in c:
        if any(x in c for x in ['go.id', 'or.id', '.org', 'kementerian', 'bumn']):
            return 0.5
        return 0.7
    return 0.6 


def extract_username_from_url(url):
    """Extract @username dari URL Instagram."""
    if not url:
        return ''
    m = re.search(r'instagram\.com/([^/?]+)', str(url))
    return m.group(1).strip('/') if m else str(url).strip()

def parse_kol_free():
    if not os.path.exists(KOL_FREE_IN):
        print(f'[WARN] {KOL_FREE_IN} tidak ditemukan, skip.')
        return []

    wb  = openpyxl.load_workbook(KOL_FREE_IN, read_only=True)
    ws  = wb['Sheet6']
    records = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row = list(row) + [None] * max(0, 6 - len(row))
        no_val, username, social_media, followers_raw, general_brief, contact = row[:6]

        if not username or str(username).strip() in ('', 'nan'):
            continue

        username_clean = str(username).strip()
        followers_num  = parse_followers(followers_raw)
        contact_str    = str(contact).strip() if contact else ''
        approach_score = approachability_score(contact_str)

        records.append({
            'id':              i + 1,
            'username':        username_clean,
            'social_media':    str(social_media).strip() if social_media else 'Instagram',
            'followers_raw':   str(followers_raw).strip() if followers_raw else '0',
            'followers_num':   followers_num,
            'category':        str(general_brief).strip() if general_brief else '',
            'contact':         contact_str,
            'approachability': approach_score,
            'pool_type':       'kol_free',   
            'rate_min':        0,
            'rate_max':        0,
        })

    print(f'[OK] KOLHomeless free: {len(records)} entri')
    return records

def parse_community():
    if not os.path.exists(COMM_IN):
        print(f'[WARN] {COMM_IN} tidak ditemukan, skip.')
        return []

    wb = openpyxl.load_workbook(COMM_IN, read_only=True)
    ws = wb['Sheet7']
    records = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row = list(row) + [None] * max(0, 5 - len(row))
        no_val, nama_komunitas, number, email, social_url = row[:5]

        if not nama_komunitas or str(nama_komunitas).strip() in ('', 'nan'):
            continue

        nama = str(nama_komunitas).strip()

        username = extract_username_from_url(social_url) or nama

        number_str = str(number).strip() if number else ''
        email_str  = str(email).strip()  if email  else ''
        if number_str and number_str not in ('nan', 'None', ''):
            best_contact = number_str
        elif email_str and email_str not in ('nan', 'None', ''):
            best_contact = email_str
        elif social_url:
            best_contact = str(social_url).strip()
        else:
            best_contact = 'dm'

        approach_score = approachability_score(best_contact)

        nama_lower = nama.lower()
        if any(k in nama_lower for k in ['fintech', 'keuangan', 'finansial', 'finance', 'cfo']):
            category = 'Finance & Fintech'
        elif any(k in nama_lower for k in ['umkm', 'pengusaha', 'bisnis', 'usaha', 'smesco', 'apindo', 'hipmi', 'kpmi']):
            category = 'UMKM & Entrepreneurship'
        elif any(k in nama_lower for k in ['perempuan', 'wanita', 'ibu', 'mom']):
            category = 'Women & Community'
        elif any(k in nama_lower for k in ['mahasiswa', 'pelajar', 'pemuda', 'muda']):
            category = 'Youth & Student'
        elif any(k in nama_lower for k in ['media', 'informasi', 'berita', 'news']):
            category = 'Media & Information'
        else:
            category = 'Community & Networking'

        records.append({
            'id':              i + 1,
            'username':        username,
            'nama_komunitas':  nama,
            'social_url':      str(social_url).strip() if social_url else '',
            'number':          number_str,
            'email':           email_str,
            'contact':         best_contact,
            'approachability': approach_score,
            'category':        category,
            'pool_type':       'community',
            'followers_num':   0,   
            'rate_min':        0,
            'rate_max':        0,
        })

    print(f'[OK] Community: {len(records)} entri')
    return records


# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)

    kol_free = parse_kol_free()
    community = parse_community()

    with open(KOL_FREE_OUT, 'w', encoding='utf-8') as f:
        json.dump(kol_free, f, ensure_ascii=False, indent=2)
    print(f'[OK] Saved kol_homeless_free.json ({len(kol_free)} entri)')

    with open(COMM_OUT, 'w', encoding='utf-8') as f:
        json.dump(community, f, ensure_ascii=False, indent=2)
    print(f'[OK] Saved community_pool.json ({len(community)} entri)')