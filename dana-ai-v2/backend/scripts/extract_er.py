import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import openpyxl, re, json, os

INSIGHT_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'insight.xlsx')
ER_OUT_PATH  = os.path.join(os.path.dirname(__file__), '..', 'data', 'er_data.json')

def extract():
    wb = openpyxl.load_workbook(INSIGHT_PATH, data_only=True)
    all_kol_er = {}

    # Sheet3: individual KOL ER tracking
    if 'Sheet3' in wb.sheetnames:
        ws = wb['Sheet3']
        for row in ws.iter_rows(values_only=True):
            if not row[1] or not isinstance(row[1], str): continue
            uname = row[1].strip().lower().replace(' ','')
            if uname in ('accountusername','') or 'total' in uname.lower(): continue
            reach    = row[4] if isinstance(row[4], (int,float)) and row[4] > 0 else 0
            total_eng= row[12] if isinstance(row[12], (int,float)) else 0
            er       = row[13] if isinstance(row[13], (int,float)) and 0 < row[13] < 200 else None
            if uname not in all_kol_er: all_kol_er[uname] = []
            all_kol_er[uname].append({'reach': reach, 'total_eng': total_eng, 'er': er})

    # Sheet4: campaign KOL performance
    if 'Sheet4' in wb.sheetnames:
        ws = wb['Sheet4']
        for row in ws.iter_rows(values_only=True):
            p = row[0]
            if not p or not isinstance(p, str): continue
            m = re.search(r'KOL\s*\d+\s*[-]\s*(?:@)?([\w.]+)', p, re.IGNORECASE)
            if m:
                uname    = m.group(1).lower()
                reach    = row[6] if isinstance(row[6], (int,float)) and row[6] > 0 else 0
                total_eng= row[14] if isinstance(row[14], (int,float)) else 0
                er       = row[15] if isinstance(row[15], (int,float)) and 0 < row[15] < 200 else None
                if uname not in all_kol_er: all_kol_er[uname] = []
                all_kol_er[uname].append({'reach': reach, 'total_eng': total_eng, 'er': er})

    # Aggregate
    summary = {}
    for uname, records in all_kol_er.items():
        if uname == 'accountusername': continue
        valid_er  = [r['er'] for r in records if r['er']]
        avg_er    = sum(valid_er)/len(valid_er) if valid_er else None
        avg_reach = sum(r['reach'] for r in records)/len(records)
        summary[uname] = {
            'avg_er_pct':  round(avg_er, 2) if avg_er else None,
            'avg_reach':   round(avg_reach),
            'post_count':  len(records),
        }

    with open(ER_OUT_PATH, 'w') as f:
        json.dump(summary, f, indent=2)

    print(f"[OK] ER data extracted: {len(summary)} KOL")
    for k, v in summary.items():
        if v['avg_er_pct']:
            print(f"   @{k}: ER={v['avg_er_pct']}%, reach={v['avg_reach']}, posts={v['post_count']}")

if __name__ == '__main__':
    extract()
